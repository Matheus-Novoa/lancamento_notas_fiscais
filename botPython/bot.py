from botcity.core import DesktopBot
from botcity.maestro import *
from openpyxl import load_workbook
import pickle
from pathlib import Path
import pygetwindow as gw
import pyautogui as gui
BotMaestroSDK.RAISE_NOT_CONNECTED = False

def main():
    bot = DesktopBot()

    pasta_dados = r'C:\Users\novoa\OneDrive\Área de Trabalho\notas_MB\planilhas\zona_sul\escola_canadenseZS_abr24'
    arquivo_planilha = Path(pasta_dados) / 'abril2024.xlsx'
    planilha = load_workbook(arquivo_planilha)['dados']
    arquivo_salvo = Path(pasta_dados) / 'dados_salvos.pkl'
    data_emissao = '30042024'
    codigo_refeicao = '3103' # zona sul
    # codigo_refeicao = '3004' # zona norte
    codigo_extra = '3609'

    def percorrer_planilha(dados):
        yield from dados

    if not arquivo_salvo.exists():
        linhas = [linha for linha in planilha.iter_rows(min_row=2, values_only=True)]
    else:
        with open(arquivo_salvo, 'rb') as f:
            linhas = pickle.load(f)

    dados_restantes = percorrer_planilha(linhas)

    titulo_janela = "Lista de Programas"
    # Obtém todas as janelas com o título especificado
    janelas = gw.getWindowsWithTitle(titulo_janela)
    # Verifica se foi encontrada alguma janela com o título especificado
    if janelas:
        # Seleciona a primeira janela encontrada (você pode iterar sobre a lista para selecionar a janela desejada)
        janela = janelas[0] 
        # Foca na janela (torna-a ativa)
        janela.activate()
    else:
        print("Nenhuma janela encontrada com o título especificado.")
    bot.wait(1000)
    gui.hotkey('alt','esc')

    for linha in dados_restantes:
        if linha[0]:
            aluno = linha[0].split()[0]
            responsavel_completo = linha[5].split()
            responsavel = ''.join([responsavel_completo[0], ' ', responsavel_completo[-1]])
            turma = linha[1]
            if 'Year' in turma or 'Y1' in turma:
                acumulador = '2'
            else: 
                acumulador = '1'
            valor = str('{:.2f}'.format(linha[3])).replace('.',',')
            refeicao = str(linha[4]).replace('.',',')
            valor_total = str('{:.2f}'.format(linha[2])).replace('.',',')
            extra = str('{:.2f}'.format(linha[-2])).replace('.',',') if linha[-2] else None
            numero_nota = str(linha[-1])
        else:
            break

        if not bot.find("ativar_edicao", matching=0.97, waiting_time=10000):
            not_found("ativar_edicao")
        bot.click()

        if not bot.find("campo_emissao", matching=0.97, waiting_time=10000):
            not_found("campo_emissao")
        bot.click_relative(104, 7, wait_after=1500, clicks=3)
        # bot.type_keys(["ctrl", "type_left"])
        # bot.type_keys(["ctrl", "shift", "type_right"])
        bot.type_key(data_emissao)

        if not bot.find("campo_acumulador", matching=0.97, waiting_time=10000):
            not_found("campo_acumulador")
        bot.click_relative(106, 5)
        bot.type_key(acumulador)
        bot.enter()
        
        if not bot.find("aba_itens", matching=0.97, waiting_time=10000):
            not_found("aba_itens")
        bot.click()
        bot.enter()
        bot.type_key('1')
        bot.enter()
        bot.type_key('1')
        bot.enter()
        bot.type_key(valor_total)
        bot.enter()

        if not bot.find("aba_contabilidade", matching=0.97, waiting_time=10000):
            not_found("aba_contabilidade")
        bot.click()
        if not bot.find("campo_valor_historico", matching=0.97, waiting_time=10000):
            not_found("campo_valor_historico")
        bot.click_relative(145, 38)
        bot.type_key(valor)
        bot.tab(presses=2)
        bot.key_end()
        vlr_provisao = f'nf 2024/{numero_nota} {responsavel} aluno {aluno}'
        bot.paste(vlr_provisao)
        bot.page_down()
        bot.key_end()
        vlr_devido = f'issqn cf {vlr_provisao}'
        bot.paste(vlr_devido)
        bot.enter()

        if float(refeicao.replace(',','.')) != 0.0:
            # if not bot.find("botao_novo", matching=0.97, waiting_time=10000):
            #     not_found("botao_novo")
            # bot.click()
            bot.enter()
            bot.type_key('100')
            bot.enter()
            bot.type_key(codigo_refeicao)
            bot.enter()
            bot.type_key(refeicao)
            bot.tab()
            bot.type_key('8')
            bot.tab()
            bot.key_end()
            bot.space()
            vlr_provisao_refeicao = f'refeição cf {vlr_provisao}'
            bot.paste(vlr_provisao_refeicao)
            bot.enter()

        if extra:
            bot.enter()
            bot.type_key('100')
            bot.enter()
            bot.type_key(codigo_extra)
            bot.enter()
            bot.type_key(extra)
            bot.tab()
            bot.type_key('8')
            bot.tab()
            bot.key_end()
            bot.space()
            vlr_provisao_extra = f'extra cf {vlr_provisao}'
            bot.paste(vlr_provisao_extra)
            bot.enter()

        resposta = gui.confirm(title='Os dados foram preenchidos corretamente?', buttons=['Continuar', 'Interromper']) 
        if resposta == 'Interromper':
            break

        if not bot.find("botao_gravar", matching=0.97, waiting_time=10000):
            not_found("botao_gravar")
        bot.click()
        # if not bot.find("botao_gravar", matching=0.97, waiting_time=10000):
        #     not_found("botao_gravar")
        # bot.click()

        if not bot.find("botao_proximo", matching=0.97, waiting_time=10000):
            not_found("botao_proximo")
        bot.click()
    
    with open(arquivo_salvo, 'wb') as f:
        print('Salvando progresso')
        pickle.dump(list(dados_restantes), f)    

def not_found(label):
    print(f"Element not found: {label}")


if __name__ == '__main__':
    main()