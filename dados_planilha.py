import openpyxl
import pickle
import pyautogui as gui
from pathlib import Path

arquivo_planilha = r'C:\Users\novoa\OneDrive\Área de Trabalho\notas_MB\planilhas\zona_norte\escola_canadense_dez23\Maple Bear Dez 23 só ativos .xlsx'
planilha = openpyxl.load_workbook(arquivo_planilha)['Alunos Matriculados Restantes']
# arquivo_salvo = Path('dados_salvos.pkl')

# def percorrer_planilha(dados):
#     yield from dados

# if not arquivo_salvo.exists():
#     linhas = [linha for linha in planilha.iter_rows(min_row=2, values_only=True)]
# else:
#     with open(arquivo_salvo, 'rb') as f:
#         linhas = pickle.load(f)

# dados_restantes = percorrer_planilha(linhas)

for linha in planilha.iter_rows(min_row=2, values_only=True):
    aluno = linha[1].split()[0]
    responsavel_completo = linha[2].split()
    responsavel = ''.join([responsavel_completo[0], ' ', responsavel_completo[-1]])
    turma = linha[3]
    acumulador = 2 if 'Year' in turma else 1
    valor = linha[5]
    refeicao = linha[7]
    valor_total = linha[8]

    print('aluno',aluno)
    print('responsavel',responsavel)
    print('turma', turma)
    print('acumulador', acumulador)
    print('valor',valor)
    print('refeição',refeicao)
    print('valor total', valor_total)
    print('\n')

#     resposta = gui.confirm(text='Os dados foram preenchidos corretamente?', buttons=['Continuar', 'Interromper'])

#     if resposta == 'Interromper':
#         break

# with open(arquivo_salvo, 'wb') as f:
#     print('Salvando progresso')
#     pickle.dump(list(dados_restantes), f)
